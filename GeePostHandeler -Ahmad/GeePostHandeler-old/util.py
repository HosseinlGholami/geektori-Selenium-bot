import time
import openpyxl
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import unicodedata
from datetime import datetime


internet_speed=0.3

next_page_Xpath='//*[@id="root"]/div/div/div[3]/div/div[2]/div[1]/div/div/div/div[2]/div/div[3]/div'
close_consumer_Xpath='//*[@id="root"]/div/div/div[3]/div/div[2]/div[2]/div/div/div/div/div[2]/div[1]'
order_Xpath=     '//*[@id="orderModalDetails"]/div[1]/div[2]'
name_Xpath=      '//*[@id="orderModalShippingDetails"]/div[1]/div[1]/div'
Tarikh_Xpath=    '//*[@id="orderModalDetails"]/div[2]/div[2]'
ersal_Xpath=     '/html/body/div[1]/div/div/div[3]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div[1]/div/div[1]/div/div[3]/div/div/div[4]'
marsoole_Xpath=  '/html/body/div[1]/div/div/div[3]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div[1]/div/div[1]/div/div[3]/div/div/div[4]/div'
inputmarso_Xpath='/html/body/div[1]/div/div/div[3]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div[1]/div[2]/input'
confirm_Xpath=   '//*[@id="root"]/div/div/div[3]/div/div[2]/div[2]/div/div/div/div/div[2]/div[2]'



def init(driver):
    driver.get('https://geektori.ir/admin/orders')
    #loginPart
    INPUT_MAIL="farzam.mirmoeini@gmail.com"           
    PASS="AHAFA12345"
    InputMail=driver.find_element_by_id("username")
    InputMail.send_keys(INPUT_MAIL)
    InputPass=driver.find_element_by_id("password")
    InputPass.send_keys(PASS)
    InputPass.send_keys(Keys.RETURN)
    #define costomer_listapth
    COUSTOMER_XPATH='//*[@id="root"]/div/div/div[3]/div/div[2]/div[1]/div/div/div/div[1]/div[2]/div'
    Page_item=[COUSTOMER_XPATH+"["+str(i)+']' for i in range(21) ]
    #define Xcel file
    DT=datetime.now().strftime('%Y-%m-%d-%m--%H\'%M')
    return DT ,Page_item,getSefareshat()
    
    
def close_consumer(driver):
    driver.find_element_by_xpath(close_consumer_Xpath).click()  
def next_page(driver,page):
    driver.find_element_by_xpath(next_page_Xpath).click()
    return page + 1
def open_consumer(driver,Page_item,index):
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH,Page_item[index]))
        ).click()
def check_open_dar_entezar(driver,Page_item,i):
    x=Page_item[i]+'/div/div[6]/div'
    y=WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,x))
        ).text
    print(y)
    return y=='در انتظار'

def scrab_details(driver,sefareshat_list):
    time.sleep(internet_speed)
    Name=WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,name_Xpath))
            ).text
    if Name in sefareshat_list:
        print("salam")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,ersal_Xpath))
            ).click()
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,marsoole_Xpath))
            ).click()
        time.sleep(internet_speed)
        InputMarsoole=WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,inputmarso_Xpath))
            )
        InputMarsoole.send_keys(sefareshat_list[Name][0])
        time.sleep(internet_speed)
        InputMarsoole.send_keys(Keys.RETURN)
        time.sleep(internet_speed)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,confirm_Xpath))
            ).click()
        sefareshat_list[Name][1]=True
    

def getSefareshat():
    wb= openpyxl.load_workbook("./data.xlsx")
    sheet=wb[wb.sheetnames[0]]
    stop=False
    i=1
    sefareshat=dict()
    while(stop!=True):
        i+=1
        code=unicodedata.normalize("NFKD",sheet.cell(i, 3).value).strip()
        sefareshat.update(
            {sheet.cell(i, 12).value : [code,False] }
            )
        
        if sheet.cell(i, 2).value==None:
            stop=True
            sefareshat.pop(None)
    return sefareshat

def save_not_done(sefareshat_list):
    wb=openpyxl.Workbook()
    DT=datetime.now().strftime('notDone-%Y-%m-%d-%m--%H\'%M')
    i=1
    for el in sefareshat_list:
        if not(sefareshat_list[el][1]):
            esh=wb["Sheet"]
            esh.cell(row=i,column=1).value=el
            esh.cell(row=i,column=2).value=sefareshat_list[el][0]
            i+=1
            
    wb.save(DT+'.xlsx')


    