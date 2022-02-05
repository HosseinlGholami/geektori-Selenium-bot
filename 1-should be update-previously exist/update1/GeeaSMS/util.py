#https://medium.com/@erayerdin/send-sms-with-python-eab7a5854d3a
import openpyxl
from unidecode import unidecode
from pyairmore.services.messaging import MessagingService
from ipaddress import IPv4Address
from pyairmore.request import AirmoreSession

def check_function(st):
        if st[0:2]=='09':
            return True
        elif st[0:3]=='989':
            return True
        elif  st[0:4]=='9809':
            return True
        else:
            return False


def Load_data():
    sheet = openpyxl.load_workbook('data.xlsx')
    esh = sheet['sheet1']
    data_list=list()
    content= esh.cell(row=1,column=1).value
    ip     = esh.cell(row=1,column=2).value.split(':')[1]
    port   = int(esh.cell(row=1,column=3).value.split(':')[1])
    run=True
    i=1
    while(run):
        i=i+1
        Name= esh.cell(row=i,column=1).value
        phone=esh.cell(row=i,column=2).value
        try:
            if phone != None:
                phone=unidecode(phone)
                x1=phone.split('-')[0].strip()
                if check_function(x1):
                    phone=x1
                x2=phone.split('-')[1].strip()
                if check_function(x2):
                    phone=x2
        except :
            pass
        data_list.append((Name,phone))
        if Name==None:
            run=False
    data_list.pop()
    return data_list,content,ip,port

def txt(name,content):
    return 'سلام '+name+' ' +content

def Send_to_all(data_list,content.Mobile_ip,Mobile_port):
    ip = IPv4Address(Mobile_ip)
    session = AirmoreSession(ip,Mobile_port)
    if session.is_server_running:
        was_accepted = session.request_authorization()
        if was_accepted:
            print('you are connected now')
        else:
            print('you are abandoned for connection')
    else:
        print('you are not in same network')
    
    service = MessagingService(session)
    for el in data_list:
        service.send_message(el[1], txt(el[0],content))
    
    

