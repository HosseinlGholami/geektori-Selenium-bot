import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import unicodedata
from unidecode import unidecode
from datetime import datetime
import pickle

darentezar_filter_Xpath='/html/body/div/div/div[5]/div/div[1]/div[2]/div[2]/div/div[1]'
darentezar_Xpath='/html/body/div/div/div[5]/div/div[1]/div[2]/div[2]/div/div[2]/div[1]'

nextPage_Xpath='/html/body/div/div/div[5]/div/div[2]/div/div/div[2]/div[3]'

internet_speed=1

confirm_Xpath='/html/body/div/div/div[5]/div/div[3]/div/div/div/div/div[2]/div[2]'

def init(driver):
    driver.get('https://geektori.ir/admin/orders')
    with open("Geektori.ir", "rb") as fp:
            file = pickle.load(fp)
    #loginPart
    INPUT_MAIL=file[0]           
    PASS=file[1]
    InputMail=driver.find_element_by_id("username")
    InputMail.send_keys(INPUT_MAIL)
    InputPass=driver.find_element_by_id("password")
    InputPass.send_keys(PASS)
    InputPass.send_keys(Keys.RETURN)
    #define Xcel file
    DT=datetime.now().strftime('%Y-%m-%d-%m--%H\'%M')
    return DT,getSefareshat()
    
def getSefareshat():
   try:
        wb= openpyxl.load_workbook("./data.xlsx")
        sheet=wb[wb.sheetnames[0]]
        stop=False
        i=1
        sefareshat=dict()
        while(stop!=True):
            i+=1
            code=unicodedata.normalize("NFKD",sheet.cell(row=i,column=3).value).strip()
            name=sheet.cell(row=i,column=12).value.strip().replace(' ','')
            sefareshat.update(
                { unidecode(name) : [{name:code},False] }
                )
            
            if sheet.cell(row=i,column=2).value==None:
                stop=True
                sefareshat.pop(None)
   except:
        print("sefareshat done")
   return sefareshat

def save_not_done(sefareshat_list):
    wb=openpyxl.Workbook()
    DT=datetime.now().strftime('notDone-%Y-%m-%d-%m--%H\'%M')
    i=1
    for el in sefareshat_list:
        if not(sefareshat_list[el][1]):
            esh=wb["Sheet"]
            x=list(sefareshat_list[el][0].items())[0]
            esh.cell(row=i,column=1).value=x[0]
            esh.cell(row=i,column=2).value=x[1]
            i+=1
            
    wb.save(DT+'.xlsx')

def end_page_check(driver,page):
    Name=WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,
    '''/html/body/div/div/div[5]/div/div[2]/div/div/div[2]/div[2]/span'''
                                            ))
            ).text
    if int(Name) == page:
        return True
    else:
        return False

driver = webdriver.Chrome('./chromedriver.exe')
#Xh : Excel handller ,DT:Date Time
DT,sefareshat_list=init(driver)
#local parameter
stop=False
time.sleep(internet_speed*3)
#select_Dar_entezar
WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH,darentezar_filter_Xpath))
        ).click()
time.sleep(internet_speed)
WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH,darentezar_Xpath))
        ).click()
time.sleep(internet_speed)
WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH,darentezar_filter_Xpath))
        ).click()
page=0
try:
    while(stop==False):
        page=page+1
        time.sleep(internet_speed*6)
        #check item in the 
        for x in range(1,21):
            try:
                #get_name of person
                Name=WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH,
                f'''/html/body/div/div/div[5]/div/div[2]/div/div/div[1]/div[{x+1}]/div[4]'''
                                                        ))
                        ).text
                Edited_name=Name.strip().replace(' ','')
                name="".join(unidecode(Edited_name))
            except:
                name=""
                stop=True
                break
            
            condition=[hash_name for hash_name in list(sefareshat_list.keys()) if name in hash_name]
            
            if condition:
                # open guys and select the postal code
                time.sleep(internet_speed)
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH,
                        f'''/html/body/div/div/div[5]/div/div[2]/div/div/div[1]/div[{x+1}]'''
                                                    ))
                    ).click()
                
                time.sleep(internet_speed)
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH,
                        '''/html/body/div/div/div[5]/div/div[3]/div/div/div/div/div[1]/div[1]/div/div[1]/div/div[3]/div/div/div[4]'''
                                                    ))
                    ).click()
                time.sleep(internet_speed)
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH,
                        '''/html/body/div/div/div[5]/div/div[3]/div/div/div/div/div[1]/div[1]/div/div[1]/div/div[3]/div/div/div[4]/div'''
                                                    ))
                    ).click()
                time.sleep(internet_speed)
                InputMarsoole=WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH,
                        '''/html/body/div/div/div[5]/div/div[3]/div/div/div/div/div[1]/div[2]/div/div/div/div/div[1]/div/div[2]/div/input'''
                                                    ))
                    )
                
                code=list(sefareshat_list[condition[0]][0].values())[0]
                                         
                InputMarsoole.send_keys(code)
                time.sleep(internet_speed)
                InputMarsoole.send_keys(Keys.RETURN)
                time.sleep(internet_speed)
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH,confirm_Xpath))
                    ).click()
                print(sefareshat_list[condition[0]])
                print(Name)
                sefareshat_list[condition[0]][1]=True
        if end_page_check(driver,page):
            stop=True
            break
        else:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,nextPage_Xpath))
                ).click()
except :
    print("somethings, goes wrong \n\n\n\n\n\n")
    print("please screen shot the last open page and show to hossein")
    print("please screen shot this page and show to hossein")
    
    time.sleep(1000*1000)
    print("somethings, goes wrong")

save_not_done(sefareshat_list)

driver.quit()
