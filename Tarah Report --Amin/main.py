from selenium import webdriver

import time
from datetime import datetime
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
from unidecode import unidecode
import pickle
internet_speed=0.3
                 
DasteBandi_Xpath='/html/body/div[1]/div/div[4]/div/div[1]/div[2]/div[3]'
nextPage_Xpath='/html/body/div[1]/div/div[4]/div/div[2]/div[1]/div/div[2]/div/div[3]'

def TarahName_Xpath(x):
    return f'/html/body/div[1]/div/div[4]/div/div[1]/div[2]/div[3]/div/div[2]/div[{x}]/div/label'
def TarahSelector_Xpath(x):
    return f'/html/body/div[1]/div/div[4]/div/div[1]/div[2]/div[3]/div/div[2]/div[{x}]'

def PnameXpath(x):
    return f'/html/body/div[1]/div/div[4]/div/div[2]/div[1]/div/div[1]/div[{x}]/div/div[2]/h2'
def PnumXpath(x):
    return f'/html/body/div[1]/div/div[4]/div/div[2]/div[1]/div/div[1]/div[{x}]/div/div[2]/div/div[1]/em[1]'
def PpriceXpath(x):
    return f'/html/body/div[1]/div/div[4]/div/div[2]/div[1]/div/div[1]/div[{x}]/div/div[2]/div/div[2]/em[1]'



def checkArlo(driver):
    time.sleep(10)
    #open dastenandi
    WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,DasteBandi_Xpath))
            ).click()
    #condition
    time.sleep(internet_speed)
    stop=False
    i=1
    while(stop==False):
        Ctg=WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH,TarahName_Xpath(i)))).text
        if (Ctg=='ARLO'):
            return i
            stop=True
        else:
            i=i+1

def init(driver):
    driver.get('https://geektori.ir/admin/products')
    #loginPart
        #loginPart
    with open("Geektori.ir", "rb") as fp:
            file = pickle.load(fp)
    INPUT_MAIL=file[0]
    PASS=file[1]
    InputMail=driver.find_element_by_id("username")
    InputMail.send_keys(INPUT_MAIL)
    InputPass=driver.find_element_by_id("password")
    InputPass.send_keys(PASS)
    InputPass.send_keys(Keys.RETURN)

    #define Xcel file
    DT=datetime.now().strftime('%Y-%m-%d-%m--%H\'%M')
    #Start inxex
    index=checkArlo(driver)
    return  DT , index

def ScrabPage(driver,i):
    product_name=WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH,PnameXpath(i))
                )).text
    product_number=WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH,PnumXpath(i))
                )).text
    product_price=WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH,PpriceXpath(i))
                )).text    
    return product_name , product_number, product_price
   
def Scrab_person(driver,esh,Tarah_name):
    wb=esh[Tarah_name]
    i=1
    j=1
    wb.cell(row=j,column=2).value='تعداد کل فروش'
    wb.cell(row=j,column=3).value='تعداد فروش فصل قبل'
    wb.cell(row=j,column=4).value='تعداد کل فروش فصل فعلی'
    wb.cell(row=j,column=5).value='قیمت'
    wb.cell(row=j,column=6).value='مجموع'
    stop=False
    j=2
    while(stop==False):
        try:
            product_name , product_number, product_price=ScrabPage(driver,i)

            wb.cell(row=j,column=1).value=product_name
            wb.cell(row=j,column=2).value=unidecode(product_number)
            wb.cell(row=j,column=4).value='=B'+str(j)+'-'+'C'+str(j)
            wb.cell(row=j,column=5).value=unidecode(product_price)
            wb.cell(row=j,column=6).value='=D'+str(j)+'*(E'+str(j)+'-0.7*2.995)'
            if (i ==20):
                #click to go next page
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH,nextPage_Xpath))
                    ).click()
                i=0
                time.sleep(3)
        except:
            wb.cell(row=j+1,column=5).value="مجموع تعداد"
            wb.cell(row=j+2,column=5).value="مجموع مبلغ"
            
            wb.cell(row=j+1,column=6).value=f'=SUM(D2:D{j})'
            wb.cell(row=j+2,column=6).value=f'=SUM(F2:F{j})'
            
            i=0
            j=0
            stop=True
        finally:
            i=i+1
            j=j+1



def Scrab(driver,th,DT):
  #select Tarah
    time.sleep(internet_speed)
    WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,TarahSelector_Xpath(th)))
            ).click()
  #deactive Dastebandi
    time.sleep(internet_speed)
    WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,DasteBandi_Xpath))
            ).click()
    #Scarab Tarah
    Tarah_name=WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,TarahName_Xpath(th)))
            ).text
    wb=openpyxl.Workbook()
    wb.create_sheet(Tarah_name)
    
    Scrab_person(driver,wb,Tarah_name)
    
    wb.remove_sheet(wb['Sheet'])
    wb.save('./Report/'+Tarah_name+'-'+DT+'.xlsx')

  #active Dastebandi
    time.sleep(internet_speed)
    WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,DasteBandi_Xpath))
            ).click()
  #deselect Tarah
    time.sleep(internet_speed)
    WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,TarahSelector_Xpath(th)))
            ).click()
    time.sleep(3)

def Action(driver,th,DT):
    i=th
    stop=False
    while(stop==False):
        try:
            Scrab(driver,i,DT)
        except:
            i=0
            stop=True
        finally:
            i=i+1



#main:            
driver = webdriver.Chrome('./chromedriver.exe')
DT,start_index_tarah=init(driver)


Action(driver,start_index_tarah,DT)

driver.quit()
