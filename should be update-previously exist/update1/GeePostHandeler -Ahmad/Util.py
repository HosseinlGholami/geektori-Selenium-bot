import time
import openpyxl
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import unicodedata
from unidecode import unidecode
from datetime import datetime
import pickle


internet_speed=0.3

next_page_Xpath='/html/body/div[1]/div/div[4]/div/div[2]/div[1]/div/div[2]/div[3]'

name_Xpath=      '/html/body/div[1]/div/div[4]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div[1]/div/div[1]/div/div[2]/div[1]/div'

ersal_Xpath=     '/html/body/div[1]/div/div[4]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div[1]/div/div[1]/div/div[3]/div/div/div[4]'
                 
marsoole_Xpath=  '/html/body/div[1]/div/div[4]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div[1]/div/div[1]/div/div[3]/div/div/div[4]/div'
inputmarso_Xpath='/html/body/div[1]/div/div[4]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div[1]/div/div[2]/div/input'

confirm_Xpath='/html/body/div[1]/div/div[4]/div/div[2]/div[2]/div/div/div/div/div[2]/div[2]'
Cancel_Xpath ='/html/body/div[1]/div/div[4]/div/div[2]/div[2]/div/div/div/div/div[2]/div[1]'

def init(driver):
    driver.get('https://geektori.ir/admin/orders')
    with open("Geektori.ir", "rb") as fp:
            file = pickle.load(fp)
            
    #loginPart
    INPUT_MAIL=file[0]           
    PASS=file[1]
    InputMail=driver.find_element_by_id("username")
    InputMail.send_keys(INPUT_MAIL)
    InputPass=driver.find_element_by_id("password")
    InputPass.send_keys(PASS)
    InputPass.send_keys(Keys.RETURN)
    #define costomer_listapth
    COUSTOMER_XPATH='/html/body/div[1]/div/div[4]/div/div[2]/div[1]/div/div[1]/div'
    Page_item=[COUSTOMER_XPATH+'['+str(i)+']' for i in range(1,22) ]
    #define Xcel file
    DT=datetime.now().strftime('%Y-%m-%d-%m--%H\'%M')
    return DT ,Page_item,getSefareshat()
    
    
def next_page(driver,page):
    driver.find_element_by_xpath(next_page_Xpath).click()
    return page + 1
def open_consumer(driver,Page_item,index):
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH,Page_item[index]))
        ).click()

def check_open_dar_entezar(driver,Page_item,i):
    x=Page_item[i]+'/div[6]/div'
    y=WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,x))
        ).text
    return y=='در انتظار'

def scrab_details(driver,sefareshat_list):
    time.sleep(internet_speed)
    Name=WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,name_Xpath))
            ).text
    name=unidecode(Name).replace(' ','')
    if name in sefareshat_list:
        time.sleep(internet_speed)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,ersal_Xpath))
            ).click()
        time.sleep(internet_speed)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,marsoole_Xpath))
            ).click()
        time.sleep(internet_speed)
        InputMarsoole=WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,inputmarso_Xpath))
            )
        code=list(sefareshat_list[name][0].values())[0]
                                 
        InputMarsoole.send_keys(code)
        time.sleep(internet_speed)
        InputMarsoole.send_keys(Keys.RETURN)
        time.sleep(internet_speed)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,confirm_Xpath))
            ).click()
        sefareshat_list[name][1]=True
    else:
        time.sleep(internet_speed)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,Cancel_Xpath))
            ).click()
        time.sleep(internet_speed)

def getSefareshat():
   try:
        wb= openpyxl.load_workbook("./data.xlsx")
        sheet=wb[wb.sheetnames[0]]
        stop=False
        i=1
        sefareshat=dict()
        while(stop!=True):
            i+=1
            code=unicodedata.normalize("NFKD",sheet.cell(row=i,column=3).value).strip()
            name=sheet.cell(row=i,column=12).value
            sefareshat.update(
                { unidecode(name).replace(' ','') : [{name:code},False] }
                )
            
            if sheet.cell(row=i,column=2).value==None:
                stop=True
                sefareshat.pop(None)
   except:
        print("sefareshat done")
    
   return sefareshat

def save_not_done(sefareshat_list):
    wb=openpyxl.Workbook()
    DT=datetime.now().strftime('notDone-%Y-%m-%d-%m--%H\'%M')
    i=1
    for el in sefareshat_list:
        if not(sefareshat_list[el][1]):
            esh=wb["Sheet"]
            x=list(sefareshat_list[el][0].items())[0]
            esh.cell(row=i,column=1).value=x[0]
            esh.cell(row=i,column=2).value=x[1]
            i+=1
            
    wb.save(DT+'.xlsx')
