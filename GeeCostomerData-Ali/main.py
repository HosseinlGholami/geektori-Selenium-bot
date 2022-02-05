from selenium import webdriver
import time
import openpyxl
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
from datetime import datetime

import pickle

internet_speed=1

darentezar_filter_Xpath='/html/body/div/div/div[5]/div/div[1]/div[2]/div[2]/div/div[1]'
darentezar_Xpath='/html/body/div/div/div[5]/div/div[1]/div[2]/div[2]/div/div[2]/div[1]'

def costumer_Xpath(x):
    return f'/html/body/div/div/div[5]/div/div[2]/div[1]/div/div[1]/div[{x}]'

                 
order_Xpath=     '//*[@id="orderModalDetails"]/div[1]/div[2]'
name_Xpath=      '//*[@id="orderModalShippingDetails"]/div[1]/div'
mobile_Xpath=    '//*[@id="orderModalShippingDetails"]/div[3]/div'
postalcode_Xpath='//*[@id="orderModalShippingDetails"]/div[4]/div'
address_Xpath=   '//*[@id="orderModalShippingDetails"]/div[5]/div'
tozihat_Xpath=   '//*[@id="orderModalShippingDetails"]/div[6]/div'

close_Xpath='/html/body/div/div/div[5]/div/div[3]/div/div/div/div/div[2]/div[1]'            
            
nextPage_Xpath='/html/body/div/div/div[5]/div/div[2]/div/div/div[2]/div[3]'

def init(driver):    
    driver.maximize_window()
    with open("Geektori.ir", "rb") as fp:
        file = pickle.load(fp)
    driver.get('https://geektori.ir/admin/orders')
    #loginPart
    INPUT_MAIL=file[0]
    PASS=file[1]
    InputMail=driver.find_element_by_id("username")
    InputMail.send_keys(INPUT_MAIL)
    InputPass=driver.find_element_by_id("password")
    InputPass.send_keys(PASS)
    InputPass.send_keys(Keys.RETURN)
    #define Xcel file
    DT=datetime.now().strftime('%Y-%m-%d-%m--%H\'%M')
    wb=openpyxl.Workbook()
    wb.create_sheet('sheet1',0)
    esh=wb['sheet1']
    esh.cell(row=1,column=1).value='نام و نام خانوادگی'
    esh.cell(row=1,column=2).value='شماره تماس'
    esh.cell(row=1,column=3).value='آدرس'
    esh.cell(row=1,column=4).value='کد پستی'
    esh.cell(row=1,column=5).value='توضیحات'
    return wb , DT 

def scrab_details(driver):
    time.sleep(internet_speed)

    Address=WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,address_Xpath))
            ).text
    Mobile=WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,mobile_Xpath))
            ).text
    Name=WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,name_Xpath))
            ).text
    PostalCode=WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,postalcode_Xpath))
            ).text
    try:
        Tozihat=WebDriverWait(driver, 1).until(
                EC.presence_of_element_located((By.XPATH,tozihat_Xpath))
                ).text
    except:
        Tozihat=''
    
    return Name,Address,Mobile,PostalCode,Tozihat

def Scarb(driver,wb):
    time.sleep(internet_speed*3)
    #select_Dar_entezar
    WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,darentezar_filter_Xpath))
            ).click()
    time.sleep(internet_speed)
    WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,darentezar_Xpath))
            ).click()
    time.sleep(internet_speed)
    WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,darentezar_filter_Xpath))
            ).click()
    time.sleep(internet_speed)
    #ScrabDetail
    stop=False
    i=1
    j=0
    while(stop==False):
        i=i+1
        j=j+1
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,costumer_Xpath(i)))
                ).click()
            Name,Address,Mobile,PostalCode,Tozihat=scrab_details(driver)
            esh=wb['sheet1']
            esh.cell(row=j+1,column=1).value=Name
            esh.cell(row=j+1,column=2).value=Mobile
            esh.cell(row=j+1,column=3).value=Address
            esh.cell(row=j+1,column=4).value=PostalCode
            esh.cell(row=j+1,column=5).value=Tozihat
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,close_Xpath))
                ).click()
        except:
            stop=True
        if i==21:
            i=1
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,nextPage_Xpath))
                ).click()

driver = webdriver.Chrome('./chromedriver.exe')
Wb,DT=init(driver)
Scarb(driver,Wb)
driver.quit()
Wb.save(DT+'.xlsx')
